"""Generate Excel comparing current vs proposed visit routes.

Growth-oriented: mirrors the bot's calculate_suggested_order logic including
segment multipliers, coverage targets, merma filters, and min order rules.
Zero-sales clients get prospecting visits with coverage-based expected orders.

Outputs: /tmp/route_optimization.xlsx
"""
import sys, os, math
from datetime import date, timedelta, datetime, timezone
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from agent_tasty.config import (
    SALESREP_ROSTER, SKU_CATALOG, FALLBACK_PRICES,
    SEGMENT_CONFIG, UT_SEGMENT_CONFIG,
    DT_COVERAGE_TARGETS, UT_COVERAGE_TARGETS,
    CANASTA_MAPPING,
)
from agent_tasty.mssql import get_mssql_connection
from agent_tasty.db import (
    SessionLocal, SalesCacheRow, ClientMermaCache, ProductPriceCache,
    compute_rfm_segment,
)
from agent_tasty.handy import get_visit_history, compute_visit_interval

from sqlalchemy import func, and_, or_

# ---------------------------------------------------------------------------
SHELF_LIFE = 18
MAX_DAILY_VISITS = 18
WORKING_DAYS_PER_WEEK = 5
DROP_COST = 7.0
GROSS_MARGIN = 0.60
OVERALL_DEVOL = 0.138
PASITAS_CODE = "PT00019"

ACTIVE_ROUTES = [
    r for r in SALESREP_ROSTER
    if r.get("handy_user_id") and not r.get("supervisor") and r["salesrep_id"] != "TEST"
]

def load_prices():
    prices = dict(FALLBACK_PRICES)
    with SessionLocal() as s:
        for row in s.query(ProductPriceCache).all():
            prices[row.product_code] = row.unit_price_usd
    return prices

PRICES = load_prices()
TRACKED_SKUS = {sku["code"]: sku for sku in SKU_CATALOG}


def get_all_clients_for_route(salesrep_id: str) -> list[dict]:
    conn = get_mssql_connection()
    try:
        c = conn.cursor()
        if salesrep_id == "DT_UT":
            c.execute(
                "SELECT Cod_Cliente, Nombre_Cliente FROM BI_CLIENTES "
                "WHERE Cod_Cliente LIKE 'DT%' OR Cod_Cliente LIKE 'UT%' "
                "ORDER BY Nombre_Cliente"
            )
        else:
            c.execute(
                "SELECT Cod_Cliente, Nombre_Cliente FROM BI_CLIENTES "
                "WHERE Vendedor = %s ORDER BY Nombre_Cliente",
                (salesrep_id,),
            )
        return [{"code": row[0], "name": row[1]} for row in c.fetchall()]
    finally:
        conn.close()


def get_client_sales_by_sku(client_code: str, days_back: int = 90) -> dict[str, dict]:
    """Per-SKU sales data using actual sale span (not inflated to days_back)."""
    cutoff = date.today() - timedelta(days=days_back)
    with SessionLocal() as session:
        rows = (
            session.query(
                SalesCacheRow.product_code,
                func.sum(SalesCacheRow.quantity).label("total_neta"),
                func.min(SalesCacheRow.sale_date).label("first_date"),
                func.max(SalesCacheRow.sale_date).label("last_date"),
                func.count(func.distinct(SalesCacheRow.sale_date)).label("sale_days"),
            )
            .filter(and_(
                SalesCacheRow.client_code == client_code,
                SalesCacheRow.sale_date >= cutoff,
                SalesCacheRow.quantity > 0,
            ))
            .group_by(SalesCacheRow.product_code)
            .all()
        )

        today = date.today()
        result = {}
        for pcode, total_neta, first_date, last_date, sale_days in rows:
            if pcode not in TRACKED_SKUS:
                continue
            total = float(total_neta) if total_neta else 0
            if total > 0 and first_date:
                # Use actual span between first and last sale, minimum 7 days
                actual_span = max((today - first_date).days, 7)
                daily_qty = total / actual_span
                result[pcode] = {
                    "daily_qty": daily_qty,
                    "total_qty": total,
                    "sale_days": sale_days,
                    "last_sale": last_date,
                }
        return result


def get_client_merma(client_code: str) -> dict[str, float]:
    with SessionLocal() as session:
        rows = session.query(ClientMermaCache).filter(
            ClientMermaCache.client_code == client_code
        ).all()
        return {r.product_code: r.merma_rate for r in rows}


def compute_proposed_order(client_code: str, sales_by_sku: dict, merma: dict,
                           visit_interval: float, rfm: dict) -> dict:
    """Mirror the bot's calculate_suggested_order logic for a proposed visit.

    Returns {product_code: suggested_qty} and total_usd.
    """
    is_dt = client_code.startswith("DT")
    is_ut = client_code.startswith("UT")
    tier = rfm["tier"]
    seg_cfg = UT_SEGMENT_CONFIG[tier] if is_ut else SEGMENT_CONFIG[tier]
    multiplier = seg_cfg["multiplier"]
    min_floor = seg_cfg["min_floor"]

    suggested = {}
    skipped = {}
    total_usd = 0.0

    for sku_code, sku_info in TRACKED_SKUS.items():
        # Skip DT-only SKUs for non-DT/UT clients
        if sku_info.get("dt_only") and not (is_dt or is_ut):
            continue

        # Merma filter: >15% = skip
        sku_merma = merma.get(sku_code, 0.0)
        if sku_merma > 0.15:
            skipped[sku_code] = f"merma {sku_merma:.0%}"
            continue

        # Get daily demand
        sku_sales = sales_by_sku.get(sku_code)
        daily = sku_sales["daily_qty"] if sku_sales else 0

        if daily > 0:
            # Apply segment multiplier + Pasitas boost
            eff_mult = multiplier * 1.25 if sku_code == PASITAS_CODE else multiplier
            # Assume shelf_stock=0 for proposed (we're planning deliveries)
            qty = max(0, math.ceil(daily * visit_interval * eff_mult))
            if min_floor > 0:
                qty = max(qty, min_floor)
        else:
            qty = 0

        # Coverage floor for DT/UT with no sales
        if qty == 0:
            if is_dt and sku_code in DT_COVERAGE_TARGETS:
                qty = DT_COVERAGE_TARGETS[sku_code]
            elif is_ut and sku_code in UT_COVERAGE_TARGETS:
                qty = UT_COVERAGE_TARGETS[sku_code]

        # Min qty floor
        if 0 < qty < 2:
            qty = 2

        if qty > 0:
            price = PRICES.get(sku_code, 0)
            # Canasta upgrade for DT
            if is_dt and sku_code in CANASTA_MAPPING and daily > 0:
                cm = CANASTA_MAPPING[sku_code]
                if qty >= cm["case_size"] - 2:
                    n_canastas = math.ceil(qty / cm["case_size"])
                    total_usd += n_canastas * cm["price_usd"]
                    suggested[sku_code] = n_canastas * cm["case_size"]
                    continue

            total_usd += qty * price
            suggested[sku_code] = qty

    return suggested, total_usd, skipped


def build_route_data():
    all_data = []

    for rep in ACTIVE_ROUTES:
        route = rep["salesrep_id"]
        salesrep_name = rep["name"]
        print(f"\n{'='*60}")
        print(f"Processing route: {route} ({salesrep_name})")
        print(f"{'='*60}")

        clients = get_all_clients_for_route(route)
        print(f"  Found {len(clients)} clients")

        for i, client in enumerate(clients):
            code = client["code"]
            name = client["name"]
            if i % 20 == 0:
                print(f"  Processing client {i+1}/{len(clients)}: {code}")

            is_dt = code.startswith("DT")
            is_ut = code.startswith("UT")

            # RFM segment
            rfm = compute_rfm_segment(code)
            tier = rfm["tier"]
            seg_cfg = UT_SEGMENT_CONFIG[tier] if is_ut else SEGMENT_CONFIG[tier]
            lookback = seg_cfg["lookback_days"]

            # Per-SKU sales (using segment-appropriate lookback)
            sales_by_sku = get_client_sales_by_sku(code, days_back=lookback)
            total_daily_usd = sum(s["daily_qty"] * PRICES.get(p, 0) for p, s in sales_by_sku.items())
            total_daily_units = sum(s["daily_qty"] for s in sales_by_sku.values())

            # Current visit interval from Handy
            try:
                visits = get_visit_history(code, days_back=90)
                current_interval = compute_visit_interval(visits)
                n_visits_90d = len(visits)
            except Exception:
                visits = []
                current_interval = 7.0
                n_visits_90d = 0

            last_visit = visits[-1] if visits else None

            # Merma rates
            merma = get_client_merma(code)
            tracked_merma = {k: v for k, v in merma.items() if k in TRACKED_SKUS}
            avg_merma = sum(tracked_merma.values()) / len(tracked_merma) if tracked_merma else 0.0

            # --- CURRENT ORDER (what they'd get today with current interval) ---
            current_suggested, current_order_usd, _ = compute_proposed_order(
                code, sales_by_sku, merma, current_interval, rfm
            )
            current_skus_ordered = sum(1 for v in current_suggested.values() if v > 0)

            # --- PROPOSED OPTIMAL INTERVAL ---
            # Goal: maximize order value per visit while respecting shelf life
            # For clients WITH sales: interval = min(shelf_life, days to accumulate strong order)
            # For clients WITHOUT sales (DT/UT): visit every shelf_life days for coverage
            # For clients WITHOUT sales (regular): visit every 2 weeks for prospecting

            min_order_usd = DROP_COST / (GROSS_MARGIN - min(avg_merma, 0.50))

            if total_daily_usd > 0:
                # Client has sales — optimize interval for order size
                # Target: order = daily * interval * multiplier should be meaningful
                # Try intervals from 3 to 18, pick the one that maximizes value while
                # staying economically viable and within shelf life
                best_interval = SHELF_LIFE
                best_order = 0
                for test_interval in range(3, SHELF_LIFE + 1):
                    _, test_usd, _ = compute_proposed_order(
                        code, sales_by_sku, merma, float(test_interval), rfm
                    )
                    if test_usd >= min_order_usd:
                        # First interval that's economically viable = most frequent viable visit
                        best_interval = test_interval
                        best_order = test_usd
                        break
                    best_order = test_usd

                optimal_interval = best_interval
            elif is_dt or is_ut:
                # DT/UT with no sales — coverage visits at shelf life interval
                optimal_interval = SHELF_LIFE
            else:
                # Regular with no sales — prospecting every 2 weeks
                optimal_interval = 14

            # Compute proposed order at optimal interval
            proposed_suggested, proposed_order_usd, proposed_skipped = compute_proposed_order(
                code, sales_by_sku, merma, float(optimal_interval), rfm
            )
            proposed_skus_ordered = sum(1 for v in proposed_suggested.values() if v > 0)
            proposed_total_units = sum(proposed_suggested.values())

            # Classification
            if total_daily_usd > 0 and proposed_order_usd >= min_order_usd:
                client_action = "OPTIMIZE"  # Has sales, profitable visit
            elif total_daily_usd > 0 and proposed_order_usd < min_order_usd:
                client_action = "CONSOLIDATE"  # Has sales but too small alone
            elif (is_dt or is_ut) and total_daily_usd == 0:
                client_action = "COVERAGE"  # DT/UT needs coverage targets
            elif total_daily_usd == 0 and n_visits_90d > 0:
                client_action = "PROSPECT"  # Was visited but no recorded sales
            else:
                client_action = "ACTIVATE"  # Never visited, no sales

            # Priority score: combines revenue potential + growth opportunity
            if total_daily_usd > 0:
                priority_score = proposed_order_usd * (1 - avg_merma)
            elif is_dt or is_ut:
                # Coverage clients get baseline priority from coverage targets
                priority_score = proposed_order_usd * 0.5  # 50% weight — speculative
            else:
                priority_score = proposed_order_usd * 0.3  # lower for unknown clients

            # Visits per month
            current_visits_month = 30.0 / current_interval if current_interval > 0 else 0
            proposed_visits_month = 30.0 / optimal_interval if optimal_interval > 0 else 0

            # Revenue projections per month
            current_monthly_rev = current_order_usd * current_visits_month
            proposed_monthly_rev = proposed_order_usd * proposed_visits_month
            monthly_delta = proposed_monthly_rev - current_monthly_rev

            row = {
                "route": route,
                "salesrep": salesrep_name,
                "client_code": code,
                "client_name": name,
                "client_type": "DT" if is_dt else "UT" if is_ut else "Regular",
                "rfm_segment": tier,
                "rfm_score": rfm["composite"],
                "action": client_action,
                # Sales metrics
                "daily_units": round(total_daily_units, 2),
                "daily_usd": round(total_daily_usd, 2),
                "avg_merma": round(avg_merma * 100, 1),
                "n_active_skus": len(sales_by_sku),
                "skus_skipped_merma": len(proposed_skipped),
                # Current
                "current_interval_days": round(current_interval, 1),
                "current_visits_month": round(current_visits_month, 1),
                "current_order_usd": round(current_order_usd, 2),
                "current_monthly_rev": round(current_monthly_rev, 2),
                "current_skus_ordered": current_skus_ordered,
                "n_visits_90d": n_visits_90d,
                "last_visit": last_visit.strftime("%Y-%m-%d") if last_visit else "N/A",
                # Proposed
                "proposed_interval_days": optimal_interval,
                "proposed_visits_month": round(proposed_visits_month, 1),
                "proposed_order_usd": round(proposed_order_usd, 2),
                "proposed_monthly_rev": round(proposed_monthly_rev, 2),
                "proposed_units_per_visit": proposed_total_units,
                "proposed_skus_ordered": proposed_skus_ordered,
                "min_economic_order": round(min_order_usd, 2),
                # Delta
                "monthly_rev_delta": round(monthly_delta, 2),
                "interval_delta": round(optimal_interval - current_interval, 1),
                "priority_score": round(priority_score, 2),
                "days_since_visit": (datetime.now(timezone.utc) - last_visit).days if last_visit else 999,
            }
            all_data.append(row)

    return all_data


def write_excel(data: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    current_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    proposed_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    delta_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
    action_fills = {
        "OPTIMIZE": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "CONSOLIDATE": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "COVERAGE": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        "PROSPECT": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "ACTIVATE": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
    }
    red_font = Font(color="CC0000", bold=True)
    green_font = Font(color="006600", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Sheet 1: Full Detail ──
    ws = wb.active
    ws.title = "Route Comparison"

    columns = [
        ("Route", 18), ("Salesrep", 16), ("Client Code", 12),
        ("Client Name", 30), ("Type", 8), ("RFM", 10), ("RFM Score", 9),
        ("Action", 13),
        ("Daily Units", 10), ("Daily USD", 10), ("Merma %", 8),
        ("Active SKUs", 10), ("SKUs Skipped (Merma)", 12),
        # Current (cols 14-20)
        ("CUR Interval (d)", 12), ("CUR Visits/Mo", 11), ("CUR $/Visit", 12),
        ("CUR $/Month", 12), ("CUR SKUs Ordered", 11),
        ("Visits (90d)", 10), ("Last Visit", 11),
        # Proposed (cols 21-28)
        ("PRO Interval (d)", 12), ("PRO Visits/Mo", 11), ("PRO $/Visit", 12),
        ("PRO $/Month", 12), ("PRO Units/Visit", 11), ("PRO SKUs Ordered", 11),
        ("Min Econ. Order $", 12),
        # Delta (cols 28-31)
        ("Monthly Rev Δ $", 13), ("Interval Δ (d)", 11),
        ("Priority Score", 12), ("Days Since Visit", 12),
    ]

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header group colors
    for col_idx in range(14, 21):
        ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        ws.cell(row=1, column=col_idx).font = header_font
    for col_idx in range(21, 28):
        ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        ws.cell(row=1, column=col_idx).font = header_font
    for col_idx in range(28, 32):
        ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws.cell(row=1, column=col_idx).font = header_font

    data.sort(key=lambda r: (r["route"], -r["priority_score"]))

    keys = [
        "route", "salesrep", "client_code", "client_name", "client_type",
        "rfm_segment", "rfm_score", "action",
        "daily_units", "daily_usd", "avg_merma", "n_active_skus", "skus_skipped_merma",
        "current_interval_days", "current_visits_month", "current_order_usd",
        "current_monthly_rev", "current_skus_ordered", "n_visits_90d", "last_visit",
        "proposed_interval_days", "proposed_visits_month", "proposed_order_usd",
        "proposed_monthly_rev", "proposed_units_per_visit", "proposed_skus_ordered",
        "min_economic_order",
        "monthly_rev_delta", "interval_delta", "priority_score", "days_since_visit",
    ]

    usd_keys = {"daily_usd", "current_order_usd", "current_monthly_rev",
                "proposed_order_usd", "proposed_monthly_rev", "min_economic_order",
                "monthly_rev_delta", "priority_score"}

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(keys, 1):
            val = row_data[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            if 14 <= col_idx <= 20:
                cell.fill = current_fill
            elif 21 <= col_idx <= 27:
                cell.fill = proposed_fill
            elif 28 <= col_idx <= 31:
                cell.fill = delta_fill

            if key in usd_keys:
                cell.number_format = '$#,##0.00'
            elif key == "avg_merma":
                cell.number_format = '0.0"%"'

            if key == "action":
                cell.fill = action_fills.get(val, PatternFill())
                cell.font = Font(bold=True)

            if key == "monthly_rev_delta":
                if val > 0:
                    cell.font = green_font
                elif val < 0:
                    cell.font = red_font

            if key == "interval_delta":
                if val > 0:
                    cell.font = green_font
                elif val < 0:
                    cell.font = red_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data)+1}"

    # ── Sheet 2: Route Summary ──
    ws2 = wb.create_sheet("Route Summary")
    summary_cols = [
        ("Route", 20), ("Salesrep", 18), ("Total Clients", 12),
        ("OPTIMIZE", 10), ("CONSOLIDATE", 12), ("COVERAGE", 10),
        ("PROSPECT", 10), ("ACTIVATE", 10),
        ("CUR $/Month", 14), ("PRO $/Month", 14), ("Rev Δ $/Month", 14),
        ("CUR Visits/Mo", 13), ("PRO Visits/Mo", 13),
        ("Capacity % (CUR)", 14), ("Capacity % (PRO)", 14),
    ]

    for col_idx, (name, width) in enumerate(summary_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    routes = defaultdict(list)
    for row_data in data:
        routes[row_data["route"]].append(row_data)

    monthly_capacity = MAX_DAILY_VISITS * WORKING_DAYS_PER_WEEK * 4.3

    for row_idx, (route, clients) in enumerate(sorted(routes.items()), 2):
        salesrep = clients[0]["salesrep"]
        total = len(clients)
        actions = defaultdict(int)
        for c in clients:
            actions[c["action"]] += 1

        cur_rev = sum(c["current_monthly_rev"] for c in clients)
        pro_rev = sum(c["proposed_monthly_rev"] for c in clients)
        cur_visits = sum(c["current_visits_month"] for c in clients)
        pro_visits = sum(c["proposed_visits_month"] for c in clients)

        vals = [
            route, salesrep, total,
            actions["OPTIMIZE"], actions["CONSOLIDATE"], actions["COVERAGE"],
            actions["PROSPECT"], actions["ACTIVATE"],
            round(cur_rev, 2), round(pro_rev, 2), round(pro_rev - cur_rev, 2),
            round(cur_visits, 1), round(pro_visits, 1),
            round(cur_visits / monthly_capacity * 100, 1),
            round(pro_visits / monthly_capacity * 100, 1),
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (9, 10, 11):
                cell.number_format = '$#,##0.00'
            if col_idx in (14, 15):
                cell.number_format = '0.0"%"'
            if col_idx == 11:
                cell.font = green_font if val > 0 else red_font if val < 0 else Font()

    ws2.freeze_panes = "A2"

    # ── Sheet 3: Proposed Weekly Plan ──
    ws3 = wb.create_sheet("Proposed Weekly Plan")
    plan_cols = [
        ("Route", 20), ("Salesrep", 18), ("Day", 12),
        ("Client Code", 12), ("Client Name", 30), ("Action", 12),
        ("RFM", 10), ("Expected Order $", 14), ("Units", 8),
        ("SKUs", 8), ("Days Overdue", 11),
    ]
    for col_idx, (name, width) in enumerate(plan_cols, 1):
        cell = ws3.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    plan_row = 2
    day_names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]

    for route, clients in sorted(routes.items()):
        salesrep = clients[0]["salesrep"]

        # Score ALL clients (including zero-sales for coverage/prospecting)
        scored = []
        for c in clients:
            interval = c["proposed_interval_days"]
            if c["proposed_order_usd"] <= 0 and c["action"] not in ("COVERAGE", "ACTIVATE"):
                continue
            urgency = c["days_since_visit"] / interval if interval > 0 else 0
            # Boost urgency for coverage/activate to ensure they get scheduled
            if c["action"] in ("COVERAGE", "ACTIVATE") and c["days_since_visit"] > interval:
                urgency *= 1.5
            scored.append((urgency, c))
        scored.sort(key=lambda x: (-x[0], -x[1]["priority_score"]))

        day_idx = 0
        day_count = 0
        daily_slots = MAX_DAILY_VISITS

        for urgency, c in scored:
            if day_idx >= 5:
                break
            ws3.cell(row=plan_row, column=1, value=route).border = thin_border
            ws3.cell(row=plan_row, column=2, value=salesrep).border = thin_border
            ws3.cell(row=plan_row, column=3, value=day_names[day_idx]).border = thin_border
            ws3.cell(row=plan_row, column=4, value=c["client_code"]).border = thin_border
            ws3.cell(row=plan_row, column=5, value=c["client_name"]).border = thin_border
            action_cell = ws3.cell(row=plan_row, column=6, value=c["action"])
            action_cell.border = thin_border
            action_cell.fill = action_fills.get(c["action"], PatternFill())
            ws3.cell(row=plan_row, column=7, value=c["rfm_segment"]).border = thin_border
            usd_cell = ws3.cell(row=plan_row, column=8, value=c["proposed_order_usd"])
            usd_cell.number_format = '$#,##0.00'
            usd_cell.border = thin_border
            ws3.cell(row=plan_row, column=9, value=c["proposed_units_per_visit"]).border = thin_border
            ws3.cell(row=plan_row, column=10, value=c["proposed_skus_ordered"]).border = thin_border
            overdue = max(0, c["days_since_visit"] - c["proposed_interval_days"])
            od_cell = ws3.cell(row=plan_row, column=11, value=overdue)
            od_cell.border = thin_border
            if overdue > 0:
                od_cell.font = red_font

            plan_row += 1
            day_count += 1
            if day_count >= daily_slots:
                day_count = 0
                day_idx += 1

    ws3.freeze_panes = "A2"

    # ── Sheet 4: Action Legend ──
    ws4 = wb.create_sheet("Legend")
    legends = [
        ("Action", "Description", "Visit Strategy"),
        ("OPTIMIZE", "Has sales, order meets minimum", "Visit at optimal interval to maximize $/visit"),
        ("CONSOLIDATE", "Has sales but order too small for solo visit", "Pair with nearby clients or extend interval"),
        ("COVERAGE", "DT/UT client, no recent sales — needs product coverage", "Visit every 18d with coverage target SKUs"),
        ("PROSPECT", "Was visited but no recorded sales in BI", "Visit every 14d to convert — potential new buyer"),
        ("ACTIVATE", "Never visited, no sales — dormant client", "Schedule first visit to assess potential"),
    ]
    for row_idx, (a, b, c) in enumerate(legends, 1):
        ws4.cell(row=row_idx, column=1, value=a).font = Font(bold=(row_idx == 1))
        ws4.cell(row=row_idx, column=2, value=b)
        ws4.cell(row=row_idx, column=3, value=c)
        if row_idx > 1 and a in action_fills:
            ws4.cell(row=row_idx, column=1).fill = action_fills[a]
    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 50
    ws4.column_dimensions["C"].width = 55

    wb.save(output_path)
    print(f"\n✅ Excel saved to: {output_path}")
    print(f"   Total rows: {len(data)} clients across {len(routes)} routes")


if __name__ == "__main__":
    output = "/tmp/route_optimization.xlsx"
    print("Building route optimization data...")
    data = build_route_data()
    print(f"\nWriting Excel ({len(data)} clients)...")
    write_excel(data, output)
